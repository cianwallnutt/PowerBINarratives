import requests
import pandas as pd
import json
import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def get_access_token():
    # Define the OAuth2 token endpoint
    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'

    # Set up the request payload
    payload = {
        'grant_type': 'client_credentials',
        'client_id': client_id,
        'client_secret': client_secret,
        'scope': 'https://analysis.windows.net/powerbi/api/.default'
    }
    # Request an access token
    response = requests.post(token_url, data=payload)
    response.raise_for_status()  # Raise an exception for HTTP errors
    # Extract the access token from the response
    return response.json().get('access_token')

def execute_dax_query(dax_query):
    access_token = get_access_token()
    # Define the endpoint URL for executing the DAX query
    execute_query_url = f'https://api.powerbi.com/v1.0/myorg/datasets/{dataset_id}/executeQueries'
    
    # Set up the headers, including the authorization header with the access token
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }
    
    # Set up the request payload for the DAX query
    query_payload = {
        "queries": [
            {
                "query": dax_query
            }
        ],
        "serializerSettings": {
            "includeNulls": True
        }
    }
    
    # Perform the POST request to execute the DAX query
    query_response = requests.post(execute_query_url, headers=headers, data=json.dumps(query_payload))
    
    # Check if the request was successful
    if query_response.status_code == 200:
        # Parse the JSON response
        query_data = query_response.json()
        return query_data
    # elif query_response.status_code == 403:
    #     access_token = get_access_token()
    #     execute_dax_query(dax_query)
    #     return
    else:
        print(f"Failed to execute DAX query: {query_response.status_code}")
        print(query_response.text)
        return None

def parse_dax_results_to_dataframe(query_data):
    if query_data and 'results' in query_data and len(query_data['results']) > 0:
        result = query_data['results'][0]
        if 'tables' in result and len(result['tables']) > 0:
            table = result['tables'][0]
            if 'rows' in table:
                rows = table['rows']
                # Extract column names from the first row's keys
                if len(rows) > 0:
                    columns = list(rows[0].keys())                
                    # Convert rows to a list of lists
                    data = [[row.get(col) for col in columns] for row in rows]
                    df = pd.DataFrame(data, columns=columns)
                    return df
    return pd.DataFrame()

def PowerBI_Query(dax_query):
    qd = execute_dax_query(dax_query)
    return parse_dax_results_to_dataframe(qd)

def HTTP_GPT(prompt, data):
    url = f"https://apim-emt-aip-prod-01.azure-api.net/openai/deployments/{deployment_id}/chat/completions?api-version={api_version}"
    # Define the headers
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "Ocp-Apim-Subscription-Key": sub_key
    }

    # Define the body
    body = {
        "stream": False,
        "stop": "None",
        "max_tokens": 4096,
        "presence_penalty": 0,
        "frequency_penalty": 0,
        "logit_bias": {},
        "user": "sustainability-scrapper-dev",
        "messages": [
            {
                "role": "system",
                "content": prompt
            },
            {
                "role": "user",
                "content": data
            }
        ]
    }
    response = requests.post(url, headers=headers, json=body)
    if response.status_code == 200: return response.json()['choices'][0]['message']['content']
    else: return f"Request failed: {response.status_code} || {response.text}"

def iterateNarratives(data,prompt,summary_prompt):
    narratives = [] 
    for d in data: narratives.append(HTTP_GPT(prompt, d))
    return HTTP_GPT(summary_prompt, json.dumps(narratives))

def narratives_init(n_tenant_id, n_client_id, n_client_secret, n_workspace_name, n_dataset_name, n_sub_key, n_deployment_id, n_api_version, n_timestamp):
    global tenant_id, client_id, client_secret, workspace_name, dataset_name, sub_key, deployment_id, api_version, dataset_id, access_token, timestamp
    tenant_id = n_tenant_id
    client_id = n_client_id
    client_secret = n_client_secret
    workspace_name = n_workspace_name
    dataset_name = n_dataset_name
    sub_key = n_sub_key
    deployment_id = n_deployment_id
    api_version = n_api_version
    timestamp = n_timestamp
    access_token = get_access_token()

    print(f"Access Token: {access_token}")

    # Define the endpoint URL for listing workspaces
    url = 'https://api.powerbi.com/v1.0/myorg/groups'

    # Set up the headers, including the authorization header with the access token
    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    # Perform the GET request to list workspaces
    response = requests.get(url, headers=headers)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the JSON response
        data = response.json()
        workspace_id = None
        for group in data['value']:
            if group['name'] == workspace_name:
                workspace_id = group['id']
                break
        
        if workspace_id:
            print(f"Workspace ID for '{workspace_name}': {workspace_id}")
            
            # Define the endpoint URL for querying datasets in the specific workspace
            datasets_url = f'https://api.powerbi.com/v1.0/myorg/groups/{workspace_id}/datasets'
            
            # Perform the GET request to list datasets in the specific workspace
            datasets_response = requests.get(datasets_url, headers=headers)
            
            # Check if the request was successful
            if datasets_response.status_code == 200:
                # Parse the JSON response
                datasets_data = datasets_response.json()
                dataset_id = None
                for dataset in datasets_data['value']:
                    if dataset['name'] == dataset_name:
                        dataset_id = dataset['id']
                        break
                
                if dataset_id:
                    print(f"Dataset ID for '{dataset_name}': {dataset_id}")
                else:
                    print(f"Dataset '{dataset_name}' not found.")
            else:
                print(f"Failed to retrieve datasets: {datasets_response.status_code}")
                print(datasets_response.text)
        else:
            print(f"Workspace '{workspace_name}' not found.")
    else:
        print(f"Failed to retrieve workspaces: {response.status_code}")
        print(response.text)

def single_level(query_json):
    concatenated_dimensions = ""
    for i, dimension in enumerate(query_json['dimensions']):
        if i == 0: 
            concatenated_dimensions = dimension
            key = f"CALCULATE(MAX({dimension}))"
        else: 
            concatenated_dimensions += "," + dimension
            key = f"CONCATENATE({key},CALCULATE(MAX({dimension})))"
    q = f"""EVALUATE
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {query_json['filters']},{query_json['metrics']},
            "Key", {key}
        )"""
    print(f"Grouping: {concatenated_dimensions}")
    out = PowerBI_Query(q)
    return json.dumps(out.to_dict(orient='records'))

def dynamic_ranking(dimensions,dimension_keys,ranked_dimension,rank_depth,drill_metric,filters,metrics):
    #Generate Primary Key and Primary Dimensions
    concatenated_dimensions = ""
    for i, dimension in enumerate(dimensions):
        if i == 0: 
            concatenated_dimensions = dimension
            primary_key = f"CALCULATE(MAX({dimension}))"
        else: 
            concatenated_dimensions += "," + dimension
            primary_key = f"CONCATENATE({primary_key},CALCULATE(MAX({dimension})))"

    ### Assign Keys if none provided 
    if len(dimension_keys) == 0:       
        q = f"""EVALUATE
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {filters},
            "Direction", IF({drill_metric}>= 0,"DESC","ASC"),
            "Key",{primary_key}
        )"""
        dimension_keys = '"' + '","'.join(PowerBI_Query(q)['[Key]'].tolist()) + '"'

    q = f"""
    EVALUATE
    VAR t_keys={{{dimension_keys}}}
    VAR toptable=
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {filters},
            "Direction", IF({drill_metric}>= 0,"DESC","ASC"),
            "Key",{primary_key}
        )
    VAR topXtable=
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {ranked_dimension},
            {filters},
            "Drill Metric", {drill_metric},
            "Key", {primary_key}
        )
    VAR rankingTable =
            FILTER(ADDCOLUMNS(NATURALINNERJOIN(topXtable, toptable),
            "RankM",
            VAR direction = [Direction]
                RETURN IF(direction = "DESC",
                    RANKX(
                        FILTER(topXtable, [Key] = EARLIER([Key])),
                        {drill_metric}
                        ,,DESC,DENSE
                    ),
                    RANKX(
                        FILTER(topXtable, [Key] = EARLIER([Key])),
                        {drill_metric}
                        ,,ASC,DENSE
                    ))),[RankM] <= {rank_depth} && [Key] in t_keys
            )
    RETURN
    CALCULATETABLE(
        SUMMARIZE(
            rankingTable,
            {concatenated_dimensions},
            {ranked_dimension},
            {metrics},
        "Key", CALCULATE(CONCATENATE({primary_key},MAX({ranked_dimension})))
        ),
        {filters}
    )"""
    return PowerBI_Query(q)

def directional_ranking(dimensions,dimension_keys,ranked_dimension,rank_depth,direction,drill_metric,filters,metrics):
    #Generate Primary Key and Primary Dimensions
    concatenated_dimensions = ""
    for i, dimension in enumerate(dimensions):
        if i == 0: 
            concatenated_dimensions = dimension
            primary_key = f"CALCULATE(MAX({dimension}))"
        else: 
            concatenated_dimensions += "," + dimension
            primary_key = f"CONCATENATE({primary_key},CALCULATE(MAX({dimension})))"

    ### Assign Keys if none provided 
    if len(dimension_keys) == 0:       
        q = f"""EVALUATE
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {filters},
            "Direction", IF({drill_metric}>= 0,"DESC","ASC"),
            "Key",{primary_key}
        )"""
        dimension_keys = '"' + '","'.join(PowerBI_Query(q)['[Key]'].tolist()) + '"'

    q = f"""
    EVALUATE
    VAR t_keys={{{dimension_keys}}}
    VAR topXtable=
        SUMMARIZECOLUMNS(
            {concatenated_dimensions},
            {ranked_dimension},
            {filters},
            "Drill Metric", {drill_metric},
            "Key", {primary_key}
        )
    VAR rankingTable=
        FILTER(
            ADDCOLUMNS(
                topXtable,
                "RankM",
                RANKX(
                    FILTER(topXtable,  [Key]= EARLIER([Key]) ),
                    {drill_metric}, ,
                    {direction},DENSE
                )
            ),
            [RankM] <= {rank_depth} && [Key] IN t_keys
    )
    RETURN
    CALCULATETABLE(
        SUMMARIZE(
            rankingTable,
            {concatenated_dimensions},
            {ranked_dimension},
            {metrics},
        "Key", CALCULATE(CONCATENATE({primary_key},MAX({ranked_dimension})))
        ),
        {filters}
    )"""
    return PowerBI_Query(q)

def iterate_dynamic(query_json,rank_depth):
    dataList = []
    dimensions = query_json['dimensions']
    dimension_keys = ""
    for i in range(len(dimensions)):
        primary_dimensions = dimensions[:i+1]
        if i + 1 < len(dimensions):
            ranked_dimension = dimensions[i + 1]
        else: break
        print(f"Grouping: {primary_dimensions} Ranking Top {rank_depth}: {ranked_dimension}")
        out = dynamic_ranking(primary_dimensions,dimension_keys,ranked_dimension,rank_depth,query_json['drill_metric'],query_json['filters'],query_json['metrics'])
        dimension_keys = '"' + '","'.join(out['[Key]'].tolist()) + '"'
        dataList.append(dimensions[i] +" by " + ranked_dimension + "\n\n" + json.dumps(out.to_dict(orient='records')))
    return dataList

def iterate_directional(query_json, rank_depth):
    dataList = []
    dimensions = query_json['dimensions']
    dimension_keys = ""
    for i in range(len(dimensions)):
        primary_dimensions = dimensions[:i+1]
        if i + 1 < len(dimensions):
            ranked_dimension = dimensions[i + 1]
        else: break
        print(f"Grouping: {primary_dimensions} Ranking Top {rank_depth}: {ranked_dimension}")
        out = directional_ranking(primary_dimensions,dimension_keys,ranked_dimension,rank_depth,query_json['direction'],query_json['drill_metric'],query_json['filters'],query_json['metrics'])
        dimension_keys = '"' + '","'.join(out['[Key]'].tolist()) + '"'
        dataList.append(dimensions[i] +" by " + ranked_dimension + "\n\n" + json.dumps(out.to_dict(orient='records')))
    return dataList

def write_dict_to_excel(file_name, dict_entry):
    # Define the fieldnames based on the dictionary keys
    fieldnames = list(dict_entry.keys())

    try:
        # Try to open the existing workbook
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        # If the file does not exist, create a new workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Narratives"
        # Write the header row
        for col_num, fieldname in enumerate(fieldnames, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}1"] = fieldname

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Write the dictionary as a row
    for col_num, fieldname in enumerate(fieldnames, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}{next_row}"] = dict_entry[fieldname]

    # Save the workbook
    workbook.save(file_name)